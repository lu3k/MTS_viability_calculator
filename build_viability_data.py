#!/usr/bin/env python3
"""
Fills in the 4 calculated data blocks that sit to the right of the raw plate
grid in this lab's MTS-assay Result sheets, given a fresh raw export straight
off the Tecan Spark.

Layout reverse-engineered from the 8 existing processed files in this folder
(MOLM/MV4 x parental/resistant x BRI_CRI/LOR_PON) and cross-checked against
the live Excel formulas still embedded in some of their cells:

  Raw plate grid (always present in a fresh export):
    - anchor cell "<>" marks the top-left of the 96-well grid.
    - the next 8 rows are well rows A-H; plate columns 1-12 run to the right
      of the anchor.
    - rows B-D, plate cols 2-11: drug 1, 10-pt dilution series, triplicate.
    - rows E-G, plate cols 2-11: drug 2, same dilution series, triplicate.
    - row F and row G, plate col 12 (M47/M48): two medium control (no-drug,
      no-cell) wells, averaged together (N51) into the background used for
      the whole plate.
    - dose axis is a fixed assay constant (TOP_DOSE_NM diluted DILUTION_FACTOR
      -fold, N_DOSES points, plus 0), not read from the plate.

  Calculated blocks this script adds (columns O onward, matching the
  existing template's cell positions exactly). Every cell in these blocks is
  written as a live Excel formula (e.g. "=L43", "=AVERAGE(M47:M48)",
  "=P43-$N$51", "=P43/$AA$44") referencing the raw plate grid, so opening the
  file in Excel shows the actual calculation and it recalculates if a raw
  value is edited:
    1. "MTS Absorbance 490 nm"                                  - raw wells
       reshaped into a 6x10 (replicate x dose) table per drug.
    2. "MTS Absorbance 490 nm Corrected"                        - block 1
       minus N51, the average of the two medium control wells.
    3. "% Viability Relative to Control (DMSO adjusted)"        - block 1
       divided by the mean of its own dose-0 (vehicle) replicates.
    4. "% Viability Relative to Control (DMSO adjusted) Corrected" - block 2
       divided by the mean of ITS dose-0 replicates.

Usage:
    python build_viability_data.py RAW.xlsx
    python build_viability_data.py RAW.xlsx -o OUT.xlsx
    python build_viability_data.py RAW.xlsx --drug1 Brigatinib --drug2 Crizotinib --cell-line1 "MOLM13 resistant"
    python build_viability_data.py RAW.xlsx --drug1 Brigatinib --cell-line1 "MOLM13 parental" --cell-line2 "MOLM13 resistant"
    python build_viability_data.py RAW.xlsx --medium-control1 0.223 --medium-control2 0.219

--drug1/--drug2 (rows B-D / rows E-G) and --cell-line1/--cell-line2 (same
row split) default to placeholder text (see PLACEHOLDER_* below) so the
labels can just be typed in by hand afterward. Rows B-D and E-G can be the
same cell line with two different drugs (the common case - pass just
--cell-line1, it's used for both), the same drug across two different cell
lines (pass the same --drug1/--drug2, different --cell-line1/--cell-line2),
or both different. The two medium control wells (M47/M48) are read from the
raw plate by default, but can be passed in explicitly with
--medium-control1/--medium-control2 (e.g. if a fresh export doesn't have
those wells measured, or to override them).
"""

import argparse
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

PLACEHOLDER_DRUG1 = "DRUG1"
PLACEHOLDER_DRUG2 = "DRUG2"
PLACEHOLDER_CELL_LINE = "CELL_LINE"

TOP_DOSE_NM = 10000.0
DILUTION_FACTOR = 2
N_DOSES = 10
DOSE_UNIT = "nM"


def find_anchor(ws):
    """Locate the '<>' cell marking the top-left corner of the raw plate grid."""
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() == "<>":
                return cell.row, cell.column
    raise ValueError("Could not find plate-grid anchor cell ('<>') in this file")


def read_raw_plate(ws, anchor_row, anchor_col):
    """8 well rows (A-H) x 12 plate columns, keyed by row letter and 1-based plate col."""
    grid = {}
    for i, letter in enumerate("ABCDEFGH"):
        r = anchor_row + 1 + i
        grid[letter] = {
            plate_col: ws.cell(row=r, column=anchor_col + plate_col).value
            for plate_col in range(1, 13)
        }
    return grid


def validate_raw_plate(grid):
    """Sanity-check a fresh export has the dose wells this template needs
    before formulas get written that would otherwise silently reference
    blanks. Medium control wells (row F/G, plate col 12) are checked
    separately in build(), since they can also be supplied on the CLI."""
    for letter in "BCDEFG":
        wells = [grid[letter][c] for c in range(2, 12)]  # plate cols 2..11
        if any(v is None for v in wells):
            raise ValueError(f"Row {letter} is missing well values in plate cols 2-11")


def col(n):
    return get_column_letter(n)


def write_dose_header_formulas(ws, row, col0):
    """col0..col0+9 = dose 0..9 ascending. Highest dose is a literal constant;
    each column to its left is that constant divided by DILUTION_FACTOR one
    more time; dose 0 is a literal 0 (mirrors X42=Y42/3 found in the source
    template, just generalized to N_DOSES/DILUTION_FACTOR/TOP_DOSE_NM)."""
    top_col = col0 + N_DOSES - 1
    ws.cell(row=row, column=top_col, value=TOP_DOSE_NM)
    for c in range(top_col - 1, col0, -1):
        ws.cell(row=row, column=c, value=f"={col(c + 1)}{row}/{DILUTION_FACTOR}")
    ws.cell(row=row, column=col0, value=0)


def build(path, drug1=None, drug2=None, cell_line1=None, cell_line2=None,
          medium_control1=None, medium_control2=None):
    wb = openpyxl.load_workbook(path)
    ws = wb.worksheets[0]

    drug1 = drug1 or PLACEHOLDER_DRUG1
    drug2 = drug2 or PLACEHOLDER_DRUG2

    # Rows B-D and E-G can be the same cell line with two different drugs
    # (the common case), the same drug across two different cell lines, or
    # both different. If only one of cell_line1/cell_line2 is given, it's
    # used for both rows (matching the old single --cell-line flag); if
    # neither is given, both fall back to the placeholder.
    if cell_line1 is None and cell_line2 is None:
        cell_line1 = cell_line2 = PLACEHOLDER_CELL_LINE
    elif cell_line1 is None:
        cell_line1 = cell_line2
    elif cell_line2 is None:
        cell_line2 = cell_line1

    # The title cell only has room for one label: collapse it to a single
    # name when both blocks share a cell line, and show "cell_line1 /
    # cell_line2" (rows B-D / rows E-G, same left-to-right order as
    # drug1/drug2) when they don't.
    cell_line_label = cell_line1 if cell_line1 == cell_line2 else f"{cell_line1} / {cell_line2}"

    anchor_row, anchor_col = find_anchor(ws)
    grid = read_raw_plate(ws, anchor_row, anchor_col)
    validate_raw_plate(grid)

    # column offsets, relative to the anchor column (normally A -> col O etc.)
    col_O = anchor_col + 14
    col_P = anchor_col + 15
    col_Z = anchor_col + 25
    col_AA = anchor_col + 26
    col_AB = anchor_col + 27
    col_AC = anchor_col + 28
    col_AM = anchor_col + 38
    col_M = anchor_col + 12   # medium control (blank) well column
    col_N = anchor_col + 13   # averaged medium control cell

    row_title1 = anchor_row - 1       # 40
    row_dose1 = anchor_row + 1        # 42
    row_data1 = anchor_row + 2        # 43: B,C,D (drug1) then E,F,G (drug2)
    row_title2 = anchor_row + 10      # 51
    row_dose2 = anchor_row + 12       # 53
    row_data2 = anchor_row + 13       # 54

    blank_row1 = anchor_row + 6       # row F, absolute (M47)
    blank_row2 = anchor_row + 7       # row G, absolute (M48)
    medium_avg_row = row_title2       # N51
    ctrl_row1_raw = row_data1 + 1     # 44 (middle of B,C,D)
    ctrl_row2_raw = row_data1 + 4     # 47 (middle of E,F,G)
    ctrl_row1_corr = row_data2 + 1    # 55
    ctrl_row2_corr = row_data2 + 4    # 58

    # Medium control (blank) wells: use the CLI value if given, otherwise
    # fall back to what's already on the raw plate. Either way the value
    # ends up written into M47/M48 so the AVERAGE formula in N51 has
    # something to reference.
    if medium_control1 is None:
        medium_control1 = grid["F"][12]
        if medium_control1 is None:
            raise ValueError(
                "Medium control well M47 (row F, plate column 12) is empty; "
                "pass --medium-control1 to supply it"
            )
    if medium_control2 is None:
        medium_control2 = grid["G"][12]
        if medium_control2 is None:
            raise ValueError(
                "Medium control well M48 (row G, plate column 12) is empty; "
                "pass --medium-control2 to supply it"
            )
    ws.cell(row=blank_row1, column=col_M, value=medium_control1)
    ws.cell(row=blank_row2, column=col_M, value=medium_control2)
    ws.cell(row=medium_avg_row, column=col_N,
            value=f"=AVERAGE({col(col_M)}{blank_row1}:{col(col_M)}{blank_row2})")

    # --- Block 1: raw absorbance, reshaped from the raw well grid ---
    ws.cell(row=row_title1, column=col_P, value="MTS Absorbance 490 nm")
    ws.cell(row=row_dose1, column=col_O, value=cell_line_label)
    write_dose_header_formulas(ws, row_dose1, col_P)
    ws.cell(row=row_dose1, column=col_Z, value=DOSE_UNIT)
    for i in range(6):  # rows 43..48
        r = row_data1 + i
        for j in range(N_DOSES):  # dose index 0..9
            raw_col = anchor_col + (11 - j)  # plate col 11-j
            ws.cell(row=r, column=col_P + j, value=f"={col(raw_col)}{r}")
    ws.cell(row=ctrl_row1_raw, column=col_O, value=drug1)
    ws.cell(row=ctrl_row1_raw, column=col_AA,
            value=f"=AVERAGE({col(col_P)}{row_data1}:{col(col_P)}{row_data1 + 2})")
    ws.cell(row=ctrl_row1_raw, column=col_AB, value=drug1)
    ws.cell(row=ctrl_row2_raw, column=col_O, value=drug2)
    ws.cell(row=ctrl_row2_raw, column=col_AA,
            value=f"=AVERAGE({col(col_P)}{row_data1 + 3}:{col(col_P)}{row_data1 + 5})")
    ws.cell(row=ctrl_row2_raw, column=col_AB, value=drug2)

    # --- Block 2: corrected absorbance = block 1 minus the averaged medium control ---
    ws.cell(row=row_title2, column=col_P, value="MTS Absorbance 490 nm Corrected")
    write_dose_header_formulas(ws, row_dose2, col_P)
    ws.cell(row=row_dose2, column=col_Z, value=DOSE_UNIT)
    for i in range(6):  # rows 54..59
        r1 = row_data1 + i
        r2 = row_data2 + i
        for j in range(N_DOSES):
            c = col_P + j
            ws.cell(row=r2, column=c,
                    value=f"={col(c)}{r1}-${col(col_N)}${medium_avg_row}")
    ws.cell(row=ctrl_row1_corr, column=col_O, value=drug1)
    ws.cell(row=ctrl_row1_corr, column=col_AA,
            value=f"=AVERAGE({col(col_P)}{row_data2}:{col(col_P)}{row_data2 + 2})")
    ws.cell(row=ctrl_row1_corr, column=col_AB, value=drug1)
    ws.cell(row=ctrl_row2_corr, column=col_O, value=drug2)
    ws.cell(row=ctrl_row2_corr, column=col_AA,
            value=f"=AVERAGE({col(col_P)}{row_data2 + 3}:{col(col_P)}{row_data2 + 5})")
    ws.cell(row=ctrl_row2_corr, column=col_AB, value=drug2)

    # --- Block 3: % viability (DMSO adjusted) = block 1 / its own control mean ---
    ws.cell(row=row_title1, column=col_AC, value="% Viability Relative to Control (DMSO adjusted)")
    write_dose_header_formulas(ws, row_dose1, col_AC)
    ws.cell(row=row_dose1, column=col_AM, value=DOSE_UNIT)
    for i in range(6):
        r = row_data1 + i
        ctrl_row = ctrl_row1_raw if i < 3 else ctrl_row2_raw
        for j in range(N_DOSES):
            src_col = col_P + j
            dst_col = col_AC + j
            ws.cell(row=r, column=dst_col,
                    value=f"={col(src_col)}{r}/${col(col_AA)}${ctrl_row}")

    # --- Block 4: % viability (DMSO adjusted) corrected = block 2 / its own control mean ---
    ws.cell(row=row_title2, column=col_AC,
            value="% Viability Relative to Control (DMSO adjusted) Corrected")
    write_dose_header_formulas(ws, row_dose2, col_AC)
    ws.cell(row=row_dose2, column=col_AM, value=DOSE_UNIT)
    for i in range(6):
        r = row_data2 + i
        ctrl_row = ctrl_row1_corr if i < 3 else ctrl_row2_corr
        for j in range(N_DOSES):
            src_col = col_P + j
            dst_col = col_AC + j
            ws.cell(row=r, column=dst_col,
                    value=f"={col(src_col)}{r}/${col(col_AA)}${ctrl_row}")

    return wb, dict(cell_line1=cell_line1, cell_line2=cell_line2, drug1=drug1, drug2=drug2,
                     medium_control1=medium_control1, medium_control2=medium_control2)


def build_directory(input_dir, output_dir=None, pattern="*.xlsx",
                     medium_control1=None, medium_control2=None):
    """Run build() over every fresh raw export in a directory.

    Cell line / drug names are left as placeholders for every file (there's
    no single override that would make sense across a whole batch of
    different plates). medium_control1/medium_control2, by contrast, sit at
    the same physical well (M47/M48) on every plate, so if given they're
    applied to every file in the batch the same way single-file mode applies
    them to one; omit them to fall back to each plate's own reading, per
    file. Already-processed outputs (from a prior run of this script, i.e.
    anything ending in "_processed.xlsx") and Excel lock files ("~$...") are
    skipped. Returns a list of (input_path, output_path, info, error)
    tuples, one per file attempted; error is None on success.
    """
    input_dir = Path(input_dir)
    output_dir = Path(output_dir) if output_dir else input_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    results = []
    for in_path in sorted(input_dir.glob(pattern)):
        if in_path.name.startswith("~$") or in_path.stem.endswith("_processed"):
            continue
        out_path = output_dir / f"{in_path.stem}_processed.xlsx"
        try:
            wb, info = build(in_path, medium_control1=medium_control1,
                              medium_control2=medium_control2)
            wb.save(out_path)
            results.append((in_path, out_path, info, None))
        except Exception as exc:
            results.append((in_path, None, None, exc))
    return results


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("input", help="fresh raw Tecan xlsx export, or a directory of them")
    ap.add_argument("-o", "--output",
                     help="single-file mode: output path (default: <input>_processed.xlsx). "
                          "directory mode: output directory (default: same as input)")
    ap.add_argument("--drug1", help="override drug 1 name (rows B-D); single-file mode only")
    ap.add_argument("--drug2", help="override drug 2 name (rows E-G); single-file mode only")
    ap.add_argument("--cell-line1", help="override cell line label for rows B-D; single-file mode "
                                          "only. If --cell-line2 is omitted, this is used for both")
    ap.add_argument("--cell-line2", help="override cell line label for rows E-G; single-file mode "
                                          "only. If --cell-line1 is omitted, this is used for both")
    ap.add_argument("--medium-control1", type=float,
                     help="medium control / blank value for M47 (row F); read from the plate if "
                          "omitted. In directory mode, applied to every file in the batch")
    ap.add_argument("--medium-control2", type=float,
                     help="medium control / blank value for M48 (row G); read from the plate if "
                          "omitted. In directory mode, applied to every file in the batch")
    args = ap.parse_args()

    in_path = Path(args.input)

    if in_path.is_dir():
        if (args.drug1 is not None or args.drug2 is not None
                or args.cell_line1 is not None or args.cell_line2 is not None):
            ap.error("--drug1/--drug2/--cell-line1/--cell-line2 apply to a single file only; run "
                      "per-file for a directory with different plates")

        results = build_directory(in_path, output_dir=args.output,
                                   medium_control1=args.medium_control1,
                                   medium_control2=args.medium_control2)
        n_ok = sum(1 for *_, err in results if err is None)
        for in_p, out_p, info, err in results:
            if err is None:
                print(f"wrote: {out_p}")
            else:
                print(f"FAILED {in_p.name}: {err}")
        print(f"{n_ok}/{len(results)} processed (calculated cells contain live Excel formulas)")
        return

    out_path = Path(args.output) if args.output else in_path.with_name(in_path.stem + "_processed.xlsx")

    wb, info = build(in_path, drug1=args.drug1, drug2=args.drug2,
                      cell_line1=args.cell_line1, cell_line2=args.cell_line2,
                      medium_control1=args.medium_control1, medium_control2=args.medium_control2)
    wb.save(out_path)

    print(f"cell line (rows B-D): {info['cell_line1']}")
    print(f"cell line (rows E-G): {info['cell_line2']}")
    print(f"drug 1 (rows B-D):    {info['drug1']}")
    print(f"drug 2 (rows E-G):    {info['drug2']}")
    print(f"medium control M47: {info['medium_control1']}")
    print(f"medium control M48: {info['medium_control2']}")
    print(f"wrote:     {out_path} (calculated cells contain live Excel formulas)")


if __name__ == "__main__":
    main()
